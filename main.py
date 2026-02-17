"""
PEP Stats Report Generator (Public Repository Version)

IMPORTANT:
SQL queries and sensitive business logic have been intentionally removed
from this public version.

To use internally:
Insert your SQL queries inside fetch_stats_for_country().
"""

import re
import json
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
from copy import copy
from datetime import datetime


# ---------------- CONFIG ----------------
PG = {
    "dbname": "your_database_name",
    "user": "user",
    "password": "password",  # Recommended: move to .env file for production
    "host": "localhost",
    "port": 5432,
}

SCHEMA = "public"
TABLE = "your_table_name"

INPUT_XLSX = r"C:\miscellaneous\stats\input.xlsx"
TEMPLATE_XLSX = r"C:\miscellaneous\stats\templet.xlsx"
OUTPUT_XLSX = r"C:\miscellaneous\stats\PEP_STATS_10022026.xlsx"

PEP_TYPE_PRIORITY = [
    "HOS", "CAB", "LEG", "MIL", "DIP", "INT",
    "GCO", "NIO", "INF", "JUD", "POL", "MUN",
    "REG", "ISO"
]

LEVEL_RE = re.compile(r"^\s*Level\s*([1-5])\s*$", re.IGNORECASE)

CREATE_INDEXES = False


# ---------------- DB CONNECTION ----------------
def connect():
    return psycopg2.connect(**PG)


def ensure_indexes(con):
    """
    Optional: Creates indexes to improve performance.
    """
    t = f"{SCHEMA}.{TABLE}"

    sqls = [
        f"""
        CREATE INDEX IF NOT EXISTS idx_pep_country_code_gin_pep
        ON {t}
        USING GIN (pep_country_code)
        WHERE is_pep = true;
        """,

        f"""
        CREATE INDEX IF NOT EXISTS idx_is_pep_subject_type
        ON {t} (subject_type)
        WHERE is_pep = true;
        """
    ]

    with con.cursor() as cur:
        for s in sqls:
            cur.execute(s)

    con.commit()


# ---------------- COUNTRY CODE PARSER ----------------
def explode_country_codes(val):
    if val is None:
        return []

    s = str(val).strip()

    if not s or s.lower() == "nan":
        return []

    if s.startswith("["):
        try:
            arr = json.loads(s)
            return [str(x).strip().upper() for x in arr if str(x).strip()]
        except Exception:
            pass

    parts = re.split(r"[,\s]+", s)

    return [p.strip().upper() for p in parts if p.strip()]


def get_all_codes_from_input(df):
    codes = []

    for v in df["Country_Code"].tolist():
        codes.extend(explode_country_codes(v))

    seen = set()
    out = []

    for c in codes:
        if c not in seen:
            seen.add(c)
            out.append(c)

    return out


# ---------------- TEMPLATE PARSER ----------------
def find_country_code_row(ws, max_scan_rows=80):

    max_col = ws.max_column

    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_col + 1):

            v = ws.cell(row=r, column=c).value

            if isinstance(v, str) and v.strip() == "Country_Code":
                return r

    raise ValueError("Country_Code header not found in template.")


def build_column_map(ws, header_row):

    sub_row = header_row + 1
    max_col = ws.max_column

    parents = {}
    last_parent = None

    for col in range(1, max_col + 1):

        v = ws.cell(row=header_row, column=col).value

        if isinstance(v, str):
            last_parent = v.strip()

        parents[col] = last_parent

    col_map = {}

    for col in range(1, max_col + 1):

        parent = parents.get(col)

        sub = ws.cell(row=sub_row, column=col).value

        if isinstance(sub, str):
            sub = sub.strip()

        if parent in PEP_TYPE_PRIORITY and isinstance(sub, str):

            m = LEVEL_RE.match(sub)

            if m:
                name = f"{parent} Level {m.group(1)}"
            else:
                name = parent

        else:
            name = parent

        if name:
            col_map.setdefault(name, []).append(col)

    return col_map


# ---------------- STYLE COPY ----------------
def copy_row_format(ws, src_row, dst_row, max_col):

    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for col in range(1, max_col + 1):

        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


# ---------------- PLACEHOLDER QUERY FUNCTION ----------------
def fetch_stats_for_country(cur, code):

    """
    PUBLIC VERSION PLACEHOLDER

    SQL queries intentionally removed.

    Insert internal SQL queries here when using in secure environment.
    """

    out = {"Country_Code": code}

    # ---------------- PLACEHOLDER ----------------
    # Example:
    #
    # sql = "YOUR QUERY HERE"
    # cur.execute(sql, (code,))
    # out.update(cur.fetchone() or {})
    #
    # --------------------------------------------

    # Role-Level matrix placeholder

    matrix = {
        f"{typ} Level {lvl}": 0
        for typ in PEP_TYPE_PRIORITY
        for lvl in range(1, 6)
    }

    out.update(matrix)

    return out


# ---------------- TEMPLATE MAPPING ----------------
def build_template_metrics(code, raw):

    out = {"Country_Code": code}

    # Placeholder mapping
    out["Total PEPs"] = raw.get("total_data", 0)

    for typ in PEP_TYPE_PRIORITY:
        for lvl in range(1, 6):
            out[f"{typ} Level {lvl}"] = raw.get(f"{typ} Level {lvl}", 0)

    return out


# ---------------- MAIN ----------------
def main():

    df = pd.read_excel(INPUT_XLSX)

    codes = get_all_codes_from_input(df)

    wb = load_workbook(TEMPLATE_XLSX)
    ws = wb.active

    header_row = find_country_code_row(ws)

    col_map = build_column_map(ws, header_row)

    country_col = col_map["Country_Code"][0]

    data_start = header_row + 2
    max_col = ws.max_column

    row_index = {}

    with connect() as con:

        if CREATE_INDEXES:
            ensure_indexes(con)

        with con.cursor(cursor_factory=RealDictCursor) as cur:

            for code in codes:

                raw = fetch_stats_for_country(cur, code)

                metrics = build_template_metrics(code, raw)

                row = ws.max_row + 1

                copy_row_format(ws, data_start, row, max_col)

                for k, v in metrics.items():

                    if k in col_map:
                        ws.cell(row=row, column=col_map[k][0]).value = v

    wb.save(OUTPUT_XLSX)

    print("Report Generated:", OUTPUT_XLSX)


# ---------------- RUN ----------------
if __name__ == "__main__":
    main()
